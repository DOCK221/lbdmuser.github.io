#!/usr/bin/env python3
"""Génère le classeur Excel de budget pour la base de données des Badges de Mickael.

Les tarifs sont saisis en euros HT, relevés en août 2026 sur les pages publiques
des fournisseurs (voir la feuille « Sources »). Mettre à jour PRIX_* puis relancer.
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SORTIE = Path(__file__).resolve().parent.parent / "budget-base-de-donnees.xlsx"

BLEU = "1F4E79"
BLEU_CLAIR = "D9E2F3"
GRIS = "F2F2F2"
VERT = "E2EFDA"
ORANGE = "FCE4D6"

TITRE = Font(bold=True, size=14, color=BLEU)
ENTETE = Font(bold=True, color="FFFFFF", size=10)
GRAS = Font(bold=True, size=10)
NORMAL = Font(size=10)
PETIT = Font(size=9, italic=True, color="595959")

FOND_ENTETE = PatternFill("solid", fgColor=BLEU)
FOND_TOTAL = PatternFill("solid", fgColor=BLEU_CLAIR)
FOND_RECO = PatternFill("solid", fgColor=VERT)
FOND_ALERTE = PatternFill("solid", fgColor=ORANGE)

HAUT = Alignment(vertical="top", wrap_text=True)
CENTRE = Alignment(horizontal="center", vertical="center", wrap_text=True)
DROITE = Alignment(horizontal="right", vertical="top")

BORDURE = Border(*(Side(style="thin", color="BFBFBF"),) * 4)

EUR = '#,##0.00\\ "€"'
EUR0 = '#,##0\\ "€"'


def ecrire_entete(ws, ligne, valeurs, largeurs=None):
    for i, valeur in enumerate(valeurs, start=1):
        cellule = ws.cell(row=ligne, column=i, value=valeur)
        cellule.font = ENTETE
        cellule.fill = FOND_ENTETE
        cellule.alignment = CENTRE
        cellule.border = BORDURE
    if largeurs:
        for i, largeur in enumerate(largeurs, start=1):
            ws.column_dimensions[get_column_letter(i)].width = largeur
    ws.row_dimensions[ligne].height = 30


def ecrire_ligne(ws, ligne, valeurs, formats=None, fond=None, gras=False):
    for i, valeur in enumerate(valeurs, start=1):
        cellule = ws.cell(row=ligne, column=i, value=valeur)
        cellule.font = GRAS if gras else NORMAL
        cellule.alignment = DROITE if isinstance(valeur, (int, float)) else HAUT
        cellule.border = BORDURE
        if fond:
            cellule.fill = fond
        if formats and formats.get(i):
            cellule.number_format = formats[i]


def feuille_synthese(wb):
    ws = wb.active
    ws.title = "Synthèse"
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 62
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Les Badges de Mickael — budget de la base de données"
    ws["A1"].font = TITRE
    ws["A2"] = "Aide aux personnes en situation de handicap · site hébergé en France (Bordeaux)"
    ws["A2"].font = PETIT

    lignes = [
        ("Contexte", ""),
        (
            "Situation actuelle",
            "Les fiches sont stockées dans le navigateur (localStorage). Elles ne sortent "
            "pas de l'appareil, ne sont pas sauvegardées et disparaissent si le navigateur "
            "est réinitialisé.",
        ),
        (
            "Besoin",
            "Une vraie base de données hébergée en France, pour que les fiches des personnes "
            "accompagnées soient partagées, sauvegardées et accessibles depuis n'importe quel "
            "appareil.",
        ),
        (
            "Nature des données",
            "Nom, identifiant, domiciliation, téléphone, aide à apporter et gestes de secours. "
            "Ce sont des données de santé et de handicap : données sensibles au sens de "
            "l'article 9 du RGPD.",
        ),
        ("", ""),
        ("Le point qui détermine le prix", ""),
        (
            "Le volume ne coûte rien",
            "Une fiche pèse environ 2 Ko. Même 1 000 personnes représentent moins de 20 Mo. "
            "La plus petite base managée du marché suffit largement, aujourd'hui et dans "
            "plusieurs années.",
        ),
        (
            "Ce qui coûte, c'est la conformité",
            "Pour des données de santé, la France impose l'hébergement chez un prestataire "
            "certifié HDS (art. L.1111-8 du code de la santé publique). C'est ce cadre, pas "
            "la taille de la base, qui fait la différence de prix.",
        ),
        (
            "À faire vérifier",
            "La certification HDS s'applique aux données de santé collectées lors d'une "
            "activité de prévention, de diagnostic, de soins ou de suivi médico-social. Selon "
            "le rôle exact de l'association, le HDS peut être obligatoire ou non. À confirmer "
            "avec un juriste ou le délégué à la protection des données avant de signer.",
        ),
        ("", ""),
        ("Les deux scénarios chiffrés", ""),
        (
            "Scénario 1 — base standard en France",
            "Base PostgreSQL managée chez un fournisseur français (Scaleway ou OVHcloud), "
            "données en France, chiffrement et sauvegardes automatiques. "
            "De 13 € à 82 € HT par mois selon le nombre de personnes.",
        ),
        (
            "Scénario 2 — base en hébergement HDS",
            "Même base, mais chez un hébergeur certifié HDS. "
            "De 50 € à 300 € HT par mois pour une petite structure. "
            "C'est le scénario à retenir si l'association relève du médico-social.",
        ),
        ("", ""),
        ("Recommandation", ""),
        (
            "Étape 1 — cadrer (avant tout achat)",
            "Faire trancher la question HDS. Réponse gratuite, et elle change le budget par "
            "un facteur 4 à 20.",
        ),
        (
            "Étape 2 — démarrer petit",
            "Base Scaleway DB-DEV-S à Paris : 11,39 € HT/mois, plus environ 1,30 € de stockage "
            "et de sauvegardes. Suffisant jusqu'à une centaine de personnes.",
        ),
        (
            "Étape 3 — passer en HDS si nécessaire",
            "Basculer vers une offre HDS quand le cadre juridique est confirmé, ou dès le "
            "départ si l'association accompagne des personnes dans un cadre médico-social.",
        ),
        ("", ""),
        ("Ce qui n'est pas dans ce budget", ""),
        (
            "Développement",
            "Remplacer le localStorage par une vraie base demande d'écrire un back-end "
            "(authentification, API, écran d'administration). C'est du développement, à chiffrer "
            "à part.",
        ),
        (
            "Hébergement du site",
            "Le site est déjà hébergé à Bordeaux. Son coût actuel n'est pas recompté ici : "
            "seule la base de données est chiffrée.",
        ),
    ]

    ligne = 4
    for titre, texte in lignes:
        if titre and not texte:
            cellule = ws.cell(row=ligne, column=1, value=titre)
            cellule.font = Font(bold=True, size=11, color=BLEU)
            ws.merge_cells(start_row=ligne, start_column=1, end_row=ligne, end_column=2)
            ws.row_dimensions[ligne].height = 22
        elif titre:
            ws.cell(row=ligne, column=1, value=titre).font = GRAS
            ws.cell(row=ligne, column=1).alignment = HAUT
            ws.cell(row=ligne, column=2, value=texte).font = NORMAL
            ws.cell(row=ligne, column=2).alignment = HAUT
            ws.row_dimensions[ligne].height = max(28, 13 * (len(texte) // 62 + 1))
        ligne += 1

    return ws


def feuille_prix(wb):
    ws = wb.create_sheet("Prix base de données")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Prix des bases de données managées — relevés août 2026, en euros HT"
    ws["A1"].font = TITRE
    ws["A2"] = (
        "Le prix « base » correspond au nœud de calcul. Le stockage et les sauvegardes "
        "sont facturés en plus, mais restent négligeables à ce volume."
    )
    ws["A2"].font = PETIT

    colonnes = [
        "Fournisseur",
        "Offre",
        "Localisation",
        "Ressources",
        "Haute disponibilité",
        "Sauvegardes",
        "Prix € HT / mois",
        "Certifié HDS",
        "Adapté pour",
    ]
    largeurs = [16, 26, 20, 20, 18, 16, 16, 22, 30]
    ecrire_entete(ws, 4, colonnes, largeurs)

    offres = [
        (
            "Scaleway",
            "Managed PostgreSQL DB-DEV-S",
            "Paris (France)",
            "2 vCPU / 2 Go",
            "Non (1 nœud)",
            "Incluses",
            11.39,
            "Offre HDS existante chez Scaleway, à confirmer sur ce type de nœud",
            "Démarrage, jusqu'à ~100 personnes",
            FOND_RECO,
        ),
        (
            "Scaleway",
            "Managed PostgreSQL DB-DEV-M",
            "Paris (France)",
            "3 vCPU / 4 Go",
            "Non (1 nœud)",
            "Incluses",
            27.89,
            "Idem",
            "100 à 300 personnes",
            None,
        ),
        (
            "Scaleway",
            "Managed PostgreSQL DB-PRO2-XXS",
            "Paris (France)",
            "2 vCPU / 8 Go",
            "Option (+~43 €)",
            "Incluses",
            80.30,
            "Idem",
            "300 à 1 000 personnes, usage intensif",
            None,
        ),
        (
            "OVHcloud",
            "PostgreSQL managé — Essential",
            "Gravelines / Roubaix (France)",
            "1 nœud",
            "Non, pas de SLA",
            "Rétention 48 h",
            54.46,
            "OVHcloud propose des offres HDS, périmètre à vérifier",
            "Alternative française à Scaleway",
            None,
        ),
        (
            "OVHcloud",
            "PostgreSQL managé — Business",
            "Gravelines / Roubaix (France)",
            "2 nœuds",
            "Oui, SLA 99,95 %",
            "Rétention 14 j",
            138.55,
            "Idem",
            "Production exigeante",
            None,
        ),
        (
            "OVHcloud",
            "PostgreSQL managé — Enterprise",
            "Gravelines / Roubaix (France)",
            "3 nœuds",
            "Oui, SLA 99,99 %",
            "Rétention 30 j",
            211.12,
            "Idem",
            "Données critiques, gros volumes",
            None,
        ),
        (
            "Hébergeur HDS",
            "Offre HDS petite structure",
            "France",
            "Selon prestataire",
            "Selon contrat",
            "Incluses",
            None,
            "Oui, c'est l'objet de l'offre",
            "Le scénario conforme pour des données de santé",
            FOND_ALERTE,
        ),
        (
            "TDF",
            "Datacenter Bordeaux-Bouliac (colocation)",
            "Bordeaux (France)",
            "Hébergement physique",
            "Équivalent Tier III",
            "À la charge du client",
            None,
            "Oui — ISO 27001 et HDS",
            "Si l'association veut rester à Bordeaux",
            None,
        ),
        (
            "Supabase",
            "Plan Pro (région Paris)",
            "Paris, société américaine",
            "Base + API incluses",
            "Selon plan",
            "Incluses",
            23.00,
            "Non",
            "Écarté : société hors UE, Cloud Act",
            None,
        ),
        (
            "VPS + PostgreSQL",
            "Serveur auto-géré",
            "France",
            "2 vCPU / 4 Go",
            "Non",
            "À faire soi-même",
            8.00,
            "Non",
            "Écarté : sécurité et sauvegardes à votre charge",
            None,
        ),
    ]

    ligne = 5
    for fournisseur, offre, lieu, res, ha, sauv, prix, hds, usage, fond in offres:
        valeurs = [fournisseur, offre, lieu, res, ha, sauv, prix, hds, usage]
        ecrire_ligne(ws, ligne, valeurs, formats={7: EUR}, fond=fond)
        if prix is None:
            cellule = ws.cell(row=ligne, column=7, value="50 à 300 €" if fournisseur == "Hébergeur HDS" else "Sur devis")
            cellule.font = NORMAL
            cellule.alignment = DROITE
            cellule.border = BORDURE
            if fond:
                cellule.fill = fond
        ws.row_dimensions[ligne].height = 30
        ligne += 1

    ligne += 1
    ws.cell(row=ligne, column=1, value="Coûts annexes de stockage (négligeables à ce volume)").font = Font(
        bold=True, size=11, color=BLEU
    )
    ligne += 1
    for libelle, prix, unite in [
        ("Stockage bloc Scaleway 5K IOPS", 0.0993, "€ HT / Go / mois"),
        ("Sauvegardes et snapshots Scaleway", 0.03, "€ HT / Go / mois"),
        ("10 Go de stockage — largement suffisant pour 1 000 fiches", 0.99, "€ HT / mois"),
        ("Nom de domaine .fr", 1.00, "€ HT / mois environ"),
        ("Certificat TLS (Let's Encrypt)", 0.00, "€ — gratuit"),
    ]:
        ws.cell(row=ligne, column=1, value=libelle).font = NORMAL
        cellule = ws.cell(row=ligne, column=2, value=prix)
        cellule.number_format = '#,##0.0000\\ "€"'
        cellule.font = NORMAL
        cellule.alignment = DROITE
        ws.cell(row=ligne, column=3, value=unite).font = PETIT
        ligne += 1

    ws.freeze_panes = "A5"
    return ws


def feuille_scenarios(wb):
    ws = wb.create_sheet("Scénarios")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Coût de la base selon le nombre de personnes accompagnées"
    ws["A1"].font = TITRE
    ws["A2"] = (
        "Hypothèse de volume : 2 Ko par fiche, arrondi à 10 Ko avec les index, l'historique "
        "et la journalisation des accès. Prix en euros HT."
    )
    ws["A2"].font = PETIT

    colonnes = [
        "Personnes",
        "Volume estimé",
        "Offre conseillée",
        "Base € / mois",
        "Stockage et sauvegardes € / mois",
        "Total standard € / mois",
        "Total standard € / an",
        "Estimation HDS € / mois",
        "Estimation HDS € / an",
        "Coût standard par personne / an",
    ]
    largeurs = [12, 14, 26, 14, 18, 16, 16, 16, 16, 18]
    ecrire_entete(ws, 4, colonnes, largeurs)

    scenarios = [
        (20, "0,2 Mo", "Scaleway DB-DEV-S", 11.39, 1.30, 60, 720),
        (50, "0,5 Mo", "Scaleway DB-DEV-S", 11.39, 1.30, 60, 720),
        (100, "1 Mo", "Scaleway DB-DEV-S", 11.39, 1.30, 80, 960),
        (200, "2 Mo", "Scaleway DB-DEV-M", 27.89, 1.30, 120, 1440),
        (500, "5 Mo", "Scaleway DB-DEV-M", 27.89, 1.60, 200, 2400),
        (1000, "10 Mo", "Scaleway DB-PRO2-XXS", 80.30, 2.00, 300, 3600),
    ]

    ligne = 5
    for personnes, volume, offre, base, annexe, hds_mois, hds_an in scenarios:
        total = base + annexe
        valeurs = [
            personnes,
            volume,
            offre,
            base,
            annexe,
            total,
            total * 12,
            hds_mois,
            hds_an,
            total * 12 / personnes,
        ]
        formats = {4: EUR, 5: EUR, 6: EUR, 7: EUR0, 8: EUR0, 9: EUR0, 10: EUR}
        ecrire_ligne(ws, ligne, valeurs, formats=formats)
        ws.row_dimensions[ligne].height = 20
        ligne += 1

    ligne += 1
    encadre = [
        "Lecture de ce tableau",
        "La base de données coûte pratiquement le même prix pour 20 ou pour 500 personnes : "
        "le volume de texte est minuscule. Ce qui fait monter la facture, c'est le nombre "
        "d'accès simultanés, la haute disponibilité et surtout le cadre HDS.",
        "La colonne « estimation HDS » reprend la fourchette de 50 à 300 € HT par mois "
        "constatée pour une petite structure. Elle est indicative : seul un devis auprès "
        "d'un hébergeur certifié donnera un chiffre ferme.",
    ]
    for texte in encadre:
        cellule = ws.cell(row=ligne, column=1, value=texte)
        cellule.font = GRAS if texte == encadre[0] else NORMAL
        cellule.alignment = HAUT
        ws.merge_cells(start_row=ligne, start_column=1, end_row=ligne, end_column=10)
        ws.row_dimensions[ligne].height = 18 if texte == encadre[0] else 32
        ligne += 1

    ws.freeze_panes = "A5"
    return ws


def feuille_frais_uniques(wb):
    ws = wb.create_sheet("Mise en place")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Frais de mise en place — à prévoir une seule fois"
    ws["A1"].font = TITRE
    ws["A2"] = "Ces montants sont des ordres de grandeur, à confirmer par devis."
    ws["A2"].font = PETIT

    ecrire_entete(ws, 4, ["Poste", "Nature", "Estimation € HT", "Commentaire"], [34, 20, 18, 52])

    postes = [
        (
            "Analyse juridique HDS et RGPD",
            "Conseil",
            "0 à 1 500 €",
            "Gratuit si l'association dispose d'un délégué à la protection des données ou "
            "passe par une permanence juridique. À faire en premier : la réponse conditionne "
            "tout le budget.",
        ),
        (
            "Développement du back-end",
            "Développement",
            "Sur devis",
            "Authentification, API, écran d'administration, journalisation des accès. "
            "C'est le poste le plus lourd et il ne dépend pas de l'hébergeur choisi.",
        ),
        (
            "Migration des fiches existantes",
            "Technique",
            "0 à 300 €",
            "Les fiches actuelles sont dans le navigateur : export en CSV ou JSON, puis import "
            "en base. Volume très faible.",
        ),
        (
            "Registre des traitements et mentions RGPD",
            "Conformité",
            "0 à 500 €",
            "Obligatoire dès qu'on traite des données de santé. Les modèles de la CNIL sont "
            "gratuits.",
        ),
        (
            "Recette et test de restauration",
            "Technique",
            "Inclus",
            "Vérifier qu'une sauvegarde se restaure vraiment. À refaire une fois par an.",
        ),
    ]

    ligne = 5
    for poste, nature, montant, commentaire in postes:
        ecrire_ligne(ws, ligne, [poste, nature, montant, commentaire])
        ws.row_dimensions[ligne].height = 44
        ligne += 1

    ws.freeze_panes = "A5"
    return ws


def feuille_sources(wb):
    ws = wb.create_sheet("Sources")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Sources et hypothèses"
    ws["A1"].font = TITRE
    ws["A2"] = "Tarifs publics relevés le 20 août 2026. Ils évoluent : à revérifier avant de signer."
    ws["A2"].font = PETIT

    ecrire_entete(ws, 4, ["Élément", "Source", "Détail"], [30, 46, 60])

    sources = [
        (
            "Prix Scaleway",
            "scaleway.com/fr/tarifs/managed-databases/",
            "DB-DEV-S à 0,0156 € HT/heure, soit 11,39 € sur 730 heures. Stockage bloc 5K à "
            "0,0993 € HT/Go/mois, sauvegardes à 0,03 € HT/Go/mois. Région Paris.",
        ),
        (
            "Prix OVHcloud",
            "ovhcloud.com/fr/public-cloud/postgresql/",
            "Essential à partir de 54,458 € HT/mois/nœud, Business 69,277 € HT/mois/nœud "
            "(2 nœuds inclus), Enterprise 70,372 € HT/mois/nœud (3 nœuds inclus).",
        ),
        (
            "Obligation HDS",
            "Article L.1111-8 du code de la santé publique",
            "Toute structure hébergeant pour le compte d'un tiers des données de santé à "
            "caractère personnel recueillies lors d'activités de prévention, de diagnostic, de "
            "soins ou de suivi médico-social doit passer par un hébergeur certifié HDS.",
        ),
        (
            "Référentiel HDS v2.0",
            "esante.gouv.fr — Agence du Numérique en Santé",
            "Seul le référentiel v2.0 est valide depuis le 16 mai 2026. Il impose un "
            "hébergement physique dans l'Espace économique européen. La liste officielle des "
            "hébergeurs certifiés est publiée par l'ANS.",
        ),
        (
            "Fourchette de prix HDS",
            "Comparatifs publics d'hébergeurs HDS, 2026",
            "50 à 300 € HT/mois pour une petite structure. Surcoût de 20 à 50 % par rapport à "
            "un hébergement classique. 600 à 1 800 € HT/mois pour un cabinet médical de 50 à "
            "100 utilisateurs avec plusieurs téraoctets.",
        ),
        (
            "Datacenter de Bordeaux",
            "tdf.fr — Data Center Bordeaux-Bouliac",
            "Certifié ISO 27001 et HDS, disponibilité annoncée 99,99 %. Pertinent si "
            "l'association souhaite garder ses données dans la région.",
        ),
        (
            "Volume de données",
            "Estimation à partir des champs du formulaire",
            "Identifiant, nom, prénom, domiciliation, téléphone, date d'inscription, aide à "
            "apporter, gestes de secours : environ 2 Ko par fiche. Arrondi à 10 Ko pour tenir "
            "compte des index, de l'historique et de la journalisation.",
        ),
        (
            "Données sensibles",
            "Article 9 du RGPD",
            "Les informations de santé et de handicap sont des données sensibles. Leur "
            "traitement demande une base légale explicite, une minimisation des données et une "
            "journalisation des accès, indépendamment de la question HDS.",
        ),
    ]

    ligne = 5
    for element, source, detail in sources:
        ecrire_ligne(ws, ligne, [element, source, detail])
        ws.row_dimensions[ligne].height = 56
        ligne += 1

    ws.freeze_panes = "A5"
    return ws


def main():
    wb = Workbook()
    feuille_synthese(wb)
    feuille_prix(wb)
    feuille_scenarios(wb)
    feuille_frais_uniques(wb)
    feuille_sources(wb)
    wb.save(SORTIE)
    print(f"Écrit : {SORTIE}")


if __name__ == "__main__":
    main()
